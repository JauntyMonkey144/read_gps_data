from flask import Flask, render_template, jsonify, request, redirect, url_for, send_file, g
from pymongo import MongoClient
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, timezone
import os
import re
import calendar
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
import secrets
import requests
from dotenv import load_dotenv

app = Flask(__name__, template_folder="templates")
CORS(app, methods=["GET", "POST"])

# ---- Init App ----
# ---- Cấu hình Domain chính cho ứng dụng ----
app.config["SERVER_NAME"] = os.environ.get("SERVER_NAME")
app.config["PREFERRED_URL_SCHEME"] = "https"

# ---- Timezone VN ----
VN_TZ = timezone(timedelta(hours=7))
load_dotenv() # Tải các biến từ file .env

# ---- MongoDB Config ----
MONGO_URI = os.getenv("MONGO_URI") 
DB_NAME = os.getenv("DB_NAME", "Sun_Database_1") 

# ---- Resend API Config ----
RESEND_API_KEY = os.getenv("RESEND_API_KEY")
RESEND_FROM_EMAIL = os.getenv("RESEND_FROM_EMAIL")

# ==============================================================================
# CẤU HÌNH KẾT NỐI MONGODB MỚI (HỖ TRỢ SERVERLESS)
# ==============================================================================

def get_db():
    """
    Tạo kết nối DB và lưu vào biến toàn cục `g` của request hiện tại.
    Chỉ kết nối khi cần thiết.
    """
    if 'db' not in g:
        g.client = MongoClient(MONGO_URI)
        g.db = g.client[DB_NAME]
    return g.db

@app.teardown_appcontext
def close_db(error):
    """
    Tự động đóng kết nối MongoDB sau khi request kết thúc.
    Điều này giúp Railway nhận diện App đang 'rảnh' để cho đi ngủ.
    """
    client = g.pop('client', None)
    if client:
        client.close()

def get_collection(name):
    """Helper để lấy collection từ kết nối hiện tại"""
    return get_db()[name]

# ==============================================================================

def send_email_resend(to_email, subject, html_body):
    """Gửi email bằng Resend API."""
    if not RESEND_API_KEY:
        print("❌ Lỗi: Biến môi trường RESEND_API_KEY chưa được đặt.")
        return False
    
    url = "https://api.resend.com/emails"
    payload = {
        "from": f"Sun Automation System <{RESEND_FROM_EMAIL}>",
        "to": [to_email],
        "subject": subject,
        "html": html_body
    }
    headers = {
        "Authorization": f"Bearer {RESEND_API_KEY}",
        "Content-Type": "application/json"
    }
    
    try:
        response = requests.post(url, json=payload, headers=headers)
        if response.status_code == 200:
            print(f"✅ Gửi email thành công đến {to_email}")
            return True
        else:
            print(f"❌ Lỗi Resend API ({response.status_code}): {response.text}")
            return False
    except Exception as e:
        print(f"❌ Lỗi ngoại lệ khi gửi email: {e}")
        return False

# ---- Trang chủ (đăng nhập chính) ----
@app.route("/")
def index():
    return render_template("index.html")

# ---- Đăng nhập API ----
@app.route("/login", methods=["POST", "GET"])
def login():
    if request.method == "GET":
        return redirect(url_for("index"))
    email = request.form.get("email")
    password = request.form.get("password")
    if not email or not password:
        return jsonify({"success": False, "message": "❌ Vui lòng nhập email và mật khẩu"}), 400
    
    # CẬP NHẬT: Dùng get_collection thay vì biến global
    admin = get_collection("admins").find_one({"email": email})
    if admin and check_password_hash(admin.get("password", ""), password):
        return jsonify({
            "success": True, "message": "✅ Đăng nhập thành công",
            "username": admin["username"], "email": admin["email"], "role": "admin"
        })
    
    user = get_collection("users").find_one({"email": email})
    if user and check_password_hash(user.get("password", ""), password):
        return jsonify({
            "success": True, "message": "✅ Đăng nhập thành công",
            "username": user["username"], "email": user["email"], "role": "user"
        })
    return jsonify({"success": False, "message": "🚫 Email hoặc mật khẩu không đúng!"}), 401

@app.route("/request-reset-password", methods=["POST"])
def request_reset_password():
    email = request.form.get("email")
    if not email:
        return """<!DOCTYPE html><html lang="vi"><head><meta charset="UTF-8"><title>Lỗi</title>...</head><body><div class="container"><p>❌ Vui lòng nhập email</p><a href="/forgot-password">Thử lại</a></div></body></html>""", 400

    # CẬP NHẬT: Dùng get_collection
    account = get_collection("admins").find_one({"email": email}) or get_collection("users").find_one({"email": email})
    if not account:
        return """<!DOCTYPE html><html lang="vi"><head><meta charset="UTF-8"><title>Lỗi</title>...</head><body><div class="container"><p>🚫 Email không tồn tại!</p><a href="/forgot-password">Thử lại</a></div></body></html>""", 404

    # Generate reset token
    token = secrets.token_urlsafe(32)
    expiration = datetime.now(VN_TZ).astimezone(timezone.utc).replace(tzinfo=None) + timedelta(days=1)
    
    # CẬP NHẬT: Dùng get_collection("reset_tokens")
    get_collection("reset_tokens").insert_one({
        "email": email,
        "token": token,
        "expiration": expiration
    })

    # Nội dung email HTML (giữ nguyên)
    reset_link = url_for("reset_password", token=token, _external=True)
    html_body = f"""
    <div style="font-family: Arial, sans-serif; color: #333; max-width: 600px; margin: auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
        <h2 style="color: #007bff; text-align: center;">Yêu cầu đặt lại mật khẩu</h2>
        <p>Xin chào,</p>
        <p>Bạn đã yêu cầu đặt lại mật khẩu cho tài khoản của mình. Vui lòng nhấp vào nút bên dưới để hoàn tất quá trình:</p>
        <p style="text-align: center; margin: 30px 0;">
            <a href="{reset_link}" 
               style="background-color: #28a745; color: white; padding: 12px 25px; text-decoration: none; border-radius: 5px; font-weight: bold; display: inline-block;">
               Đặt lại mật khẩu
            </a>
        </p>
        <p>Liên kết này sẽ hết hạn sau 1 ngày.</p>
        <p>Nếu bạn không yêu cầu đặt lại mật khẩu, vui lòng bỏ qua email này một cách an toàn.</p>
        <hr style="border: none; border-top: 1px solid #eee;">
        <p style="font-size: 0.9em; color: #6c757d; text-align: center;">
            Trân trọng,<br>
            Hệ thống tự động Sun Automation
        </p>
    </div>
    """

    if send_email_resend(email, "Yêu cầu đặt lại mật khẩu", html_body):
        return """<!DOCTYPE html><html lang="vi"><head><meta charset="UTF-8"><title>Gửi liên kết thành công</title>...</head><body><div class="container"><div class="success">✅ Email chứa liên kết đặt lại mật khẩu đã được gửi thành công! Vui lòng kiểm tra hộp thư của bạn.</div><a href="/"><button>Quay về trang chủ</button></a></div></body></html>"""
    else:
        return """<!DOCTYPE html><html lang="vi"><head><meta charset="UTF-8"><title>Lỗi</title>...</head><body><div class="container"><p>❌ Lỗi khi gửi email, vui lòng thử lại sau</p><a href="/forgot-password">Thử lại</a></div></body></html>""", 500

# ---- Trang reset mật khẩu với token (ĐÃ SỬA LỖI) ----
@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    # CẬP NHẬT: Dùng get_collection("reset_tokens")
    reset_tokens_col = get_collection("reset_tokens")

    if request.method in ["GET", "HEAD"]:
        token_data = reset_tokens_col.find_one({"token": token})
        if not token_data or token_data["expiration"] < datetime.now(timezone.utc).replace(tzinfo=None):
            return """<!DOCTYPE html><html lang="vi">...<p>🚫 Liên kết đặt lại mật khẩu không hợp lệ hoặc đã hết hạn!</p>...</html>""", 400

        return """<!DOCTYPE html><html lang="vi"><head><meta charset="UTF-8"><title>Đặt lại mật khẩu</title>...</head><body><div class="container"><h2>🔒 Đặt lại mật khẩu</h2><form method="POST"><input type="password" name="new_password" placeholder="Mật khẩu mới" required><input type="password" name="confirm_password" placeholder="Xác nhận mật khẩu" required><button type="submit">Cập nhật mật khẩu</button></form></div></body></html>"""

    if request.method == "POST":
        token_data = reset_tokens_col.find_one({"token": token})
        if not token_data or token_data["expiration"] < datetime.now(timezone.utc).replace(tzinfo=None):
            return """<!DOCTYPE html><html lang="vi">...<p>❌ Liên kết không hợp lệ hoặc đã hết hạn</p>...</html>""", 400

        new_password = request.form.get("new_password")
        confirm_password = request.form.get("confirm_password")
        if not new_password or not confirm_password:
            return """<!DOCTYPE html><html lang="vi">...<p>❌ Vui lòng điền đầy đủ thông tin</p>...</html>""".format(token), 400
        if new_password != confirm_password:
            return """<!DOCTYPE html><html lang="vi">...<p>❌ Mật khẩu xác nhận không khớp</p>...</html>""".format(token), 400

        email = token_data["email"]
        
        # CẬP NHẬT: Dùng get_collection
        admins_col = get_collection("admins")
        users_col = get_collection("users")
        
        account = admins_col.find_one({"email": email}) or users_col.find_one({"email": email})
        if not account:
            return """<!DOCTYPE html><html lang="vi">...<p>🚫 Email không tồn tại!</p>...</html>""", 404

        hashed_pw = generate_password_hash(new_password)
        collection_to_update = admins_col if "username" in account else users_col
        collection_to_update.update_one({"email": email}, {"$set": {"password": hashed_pw}})
        
        reset_tokens_col.delete_one({"token": token}) # Remove used token

        return """<!DOCTYPE html><html lang="vi"><head><title>Thay đổi mật khẩu thành công</title>...</head><body><div class="container"><div class="success">✅ Thay đổi mật khẩu thành công! Bạn có thể đăng nhập với mật khẩu mới.</div><a href="/"><button>Quay về trang chủ</button></a></div></body></html>"""

@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "GET":
        return """<!DOCTYPE html><html lang="vi"><head><meta charset="UTF-8"><title>Đặt lại mật khẩu</title>...</head><body><div class="container"><h2>🔒 Đặt lại mật khẩu</h2><form method="POST" action="/request-reset-password"><input type="email" name="email" placeholder="Email" required><button type="submit">Gửi liên kết đặt lại</button><a href="/">Quay về trang chủ</a></form></div></body></html>"""
    if request.method == "POST":
        email = request.form.get("email")
        # ... (Logic tạo tài khoản mới nếu cần, nhưng ở đây chỉ dùng request-reset-password)
        # Phần này của code gốc dường như là logic tạo mới/reset trực tiếp, 
        # nhưng bạn đã có route /request-reset-password riêng. 
        # Tôi sẽ giữ nguyên logic "đổi mật khẩu trực tiếp" nếu bạn muốn, nhưng cập nhật collection.
        
        new_password = request.form.get("new_password")
        confirm_password = request.form.get("confirm_password")
        if not all([email, new_password, confirm_password]):
             return """<!DOCTYPE html><html lang="vi">...<p>❌ Vui lòng điền đầy đủ thông tin</p>...</html>""", 400
        if new_password != confirm_password:
             return """<!DOCTYPE html><html lang="vi">...<p>❌ Mật khẩu xác nhận không khớp</p>...</html>""", 400
        
        # CẬP NHẬT: Dùng get_collection
        admins_col = get_collection("admins")
        users_col = get_collection("users")
        
        account = admins_col.find_one({"email": email}) or users_col.find_one({"email": email})
        if not account:
             return """<!DOCTYPE html><html lang="vi">...<p>🚫 Email không tồn tại!</p>...</html>""", 404
        
        hashed_pw = generate_password_hash(new_password)
        collection_to_update = admins_col if "username" in account else users_col
        collection_to_update.update_one({"email": email}, {"$set": {"password": hashed_pw}})
        return """<!DOCTYPE html><html lang="vi">...<div class="success">✅ Thay đổi mật khẩu thành công!</div>...</html>"""

# ---- Build leave query (lọc theo dateType)----
def build_leave_query(filter_type, start_date_str, end_date_str, search, date_type="CheckinTime", username=None):
    today = datetime.now(VN_TZ)
    regex_leave = re.compile("Nghỉ phép", re.IGNORECASE)
    conditions = [{"$or": [{"Tasks": regex_leave}, {"Reason": {"$exists": True}}]}]
    date_filter = {}

    start_dt, end_dt = None, None

    if filter_type == "custom" and start_date_str and end_date_str:
        try:
            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d").replace(tzinfo=VN_TZ)
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=VN_TZ)
        except ValueError:
            pass
    elif filter_type != "tất cả":
        if filter_type == "hôm nay":
            start_dt, end_dt = today.replace(hour=0, minute=0, second=0), today.replace(hour=23, minute=59, second=59)
        elif filter_type == "tuần":
            start_dt = (today - timedelta(days=today.weekday())).replace(hour=0, minute=0, second=0)
            end_dt = (start_dt + timedelta(days=6)).replace(hour=23, minute=59, second=59)
        elif filter_type == "tháng":
            start_dt = today.replace(day=1, hour=0, minute=0, second=0)
            _, last_day = calendar.monthrange(today.year, today.month)
            end_dt = today.replace(day=last_day, hour=23, minute=59, second=59)
        elif filter_type == "năm":
            start_dt = today.replace(month=1, day=1, hour=0, minute=0, second=0)
            end_dt = today.replace(month=12, day=31, hour=23, minute=59, second=59)

    if start_dt and end_dt:
        if date_type == "CheckinTime":
            date_filter = {"CreationTime": {"$gte": start_dt, "$lte": end_dt}}
        elif date_type == "ApprovalDate1":
            date_filter = {"ApprovalDate1": {"$gte": start_dt, "$lte": end_dt}}
        elif date_type == "ApprovalDate2":
            date_filter = {"ApprovalDate2": {"$gte": start_dt, "$lte": end_dt}}

    if date_filter:
        conditions.append(date_filter)

    if search:
        regex = re.compile(search, re.IGNORECASE)
        conditions.append({"$or": [{"EmployeeId": regex}, {"EmployeeName": regex}]})
    if username:
        conditions.append({"EmployeeName": username})

    return {"$and": conditions}

# ---- Build attendance query ----
def build_attendance_query(filter_type, start_date, end_date, search, username=None):
    today = datetime.now(VN_TZ)
    conditions = [{"CheckType": {"$in": ["checkin", "checkout"]}}]
    date_filter = {}
    if filter_type == "custom" and start_date and end_date:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=VN_TZ)
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=VN_TZ)
        date_filter = {"Timestamp": {"$gte": start_dt, "$lte": end_dt}}
    elif filter_type == "hôm nay":
        date_filter = {"CheckinDate": today.strftime("%d/%m/%Y")}
    elif filter_type == "tuần":
        start_dt = (today - timedelta(days=today.weekday())).replace(hour=0, minute=0, second=0)
        end_dt = (start_dt + timedelta(days=6)).replace(hour=23, minute=59, second=59)
        date_filter = {"Timestamp": {"$gte": start_dt, "$lte": end_dt}}
    elif filter_type == "tháng":
        date_filter = {"CheckinDate": {"$regex": f"/{today.month:02d}/{today.year}$"}}
    elif filter_type == "năm":
        date_filter = {"CheckinDate": {"$regex": f"/{today.year}$"}}
    if date_filter: conditions.append(date_filter)
    if search:
        regex = re.compile(search, re.IGNORECASE)
        conditions.append({"$or": [{"EmployeeId": regex}, {"EmployeeName": regex}]})
    if username:
        conditions.append({"EmployeeName": username})
    return {"$and": conditions}

# ---- Helper functions ----
def calculate_leave_days_for_month(record, export_year, export_month):
    # ... (Giữ nguyên hàm này) ...
    display_date = record.get("DisplayDate", "").strip().lower()
    start_date, end_date = None, None
    try:
        if display_date:
            if "từ" in display_date and "đến" in display_date:
                date_parts = re.findall(r"\d{4}-\d{2}-\d{2}", display_date)
                if len(date_parts) == 2:
                    start_date = datetime.strptime(date_parts[0], "%Y-%m-%d")
                    end_date = datetime.strptime(date_parts[1], "%Y-%m-%d")
            else:
                date_part = display_date.split()[0]
                start_date = end_date = datetime.strptime(date_part, "%Y-%m-%d")
        elif record.get('StartDate') and record.get('EndDate'):
             start_date = datetime.strptime(record['StartDate'], "%Y-%m-%d")
             end_date = datetime.strptime(record['EndDate'], "%Y-%m-%d")
        elif record.get('LeaveDate'):
             start_date = end_date = datetime.strptime(record['LeaveDate'], "%Y-%m-%d")
    except (ValueError, TypeError, IndexError):
        return 0.0, False

    if not start_date or not end_date:
        return 0.0, False

    days_in_month = 0.0
    _, last_day = calendar.monthrange(export_year, export_month)
    month_start = datetime(export_year, export_month, 1)
    month_end = datetime(export_year, export_month, last_day)
    if start_date == end_date:
        if month_start <= start_date <= month_end and start_date.weekday() <= 5:
            if "cả ngày" in display_date or not ("sáng" in display_date or "chiều" in display_date or record.get('Session', '').lower() in ['sáng', 'chiều']):
                days_in_month = 1.0
            else:
                days_in_month = 0.5
    else:
        current_date = max(start_date, month_start)
        end = min(end_date, month_end)
        while current_date <= end:
            if current_date.weekday() <= 5:
                days_in_month += 1.0
            current_date += timedelta(days=1)

    if days_in_month == 0:
        return 0.0, False

    status1 = str(record.get("Status1") or "").lower()
    status2 = str(record.get("Status2") or "").lower()

    if "từ chối" in status2:
        return 0.0, True
    elif "duyệt" in status2:
        return days_in_month, True
    elif "duyệt" in status1:
        return days_in_month, True
    else:
        return 0.0, True

def get_formatted_approval_date(approval_date):
    if not approval_date: return ""
    try: return approval_date.astimezone(VN_TZ).strftime("%d/%m/%Y %H:%M:%S") if isinstance(approval_date, datetime) else str(approval_date)
    except: return str(approval_date)

# ---- API lấy dữ liệu chấm công ----
@app.route("/api/attendances", methods=["GET"])
def get_attendances():
    try:
        email = request.args.get("email")
        # CẬP NHẬT: Dùng get_collection
        admins_col = get_collection("admins")
        users_col = get_collection("users")
        
        admin = admins_col.find_one({"email": email})
        user = users_col.find_one({"email": email})
        
        if not admin and not user: return jsonify({"error": "🚫 Email không tồn tại"}), 403
        username = None if admin else user["username"]
        query = build_attendance_query(
            request.args.get("filter", "hôm nay").lower(),
            request.args.get("startDate"), request.args.get("endDate"),
            request.args.get("search", "").strip(), username=username
        )
        
        # CẬP NHẬT: Dùng get_collection("alt_checkins")
        alt_checkins_col = get_collection("alt_checkins")
        all_relevant_data = list(alt_checkins_col.find(query, {"_id": 0}))
        
        daily_hours_map, monthly_hours_map = {}, {}
        emp_data = {}
        for rec in all_relevant_data:
            emp_id = rec.get("EmployeeId")
            if emp_id: emp_data.setdefault(emp_id, []).append(rec)
        
        for emp_id, records in emp_data.items():
            daily_groups = {}
            for rec in records:
                date_str = rec.get("CheckinDate")
                if date_str: daily_groups.setdefault(date_str, []).append(rec)
            
            for date_str, day_records in daily_groups.items():
                checkins = []
                for r in day_records:
                    if r.get('CheckType') == 'checkin' and r.get('Timestamp'):
                        try:
                            if isinstance(r['Timestamp'], str):
                                timestamp = datetime.strptime(r['Timestamp'], "%Y-%m-%d %H:%M:%S")
                            elif isinstance(r['Timestamp'], datetime):
                                timestamp = r['Timestamp']
                            else:
                                continue
                            checkins.append(timestamp)
                        except (ValueError, TypeError):
                            continue
                checkins = sorted(checkins)
                checkouts = []
                for r in day_records:
                    if r.get('CheckType') == 'checkout' and r.get('Timestamp'):
                        try:
                            if isinstance(r['Timestamp'], str):
                                timestamp = datetime.strptime(r['Timestamp'], "%Y-%m-%d %H:%M:%S")
                            elif isinstance(r['Timestamp'], datetime):
                                timestamp = r['Timestamp']
                            else:
                                continue
                            checkouts.append(timestamp)
                        except (ValueError, TypeError):
                            continue
                checkouts = sorted(checkouts)
                daily_seconds = 0
                if checkins and checkouts and checkouts[-1] > checkins[0]:
                    daily_seconds = (checkouts[-1] - checkins[0]).total_seconds()
                daily_hours_map[(emp_id, date_str)] = daily_seconds
                
                h, rem = divmod(daily_seconds, 3600)
                m, s = divmod(rem, 60)
                daily_hours = f"{int(h)}h {int(m)}m {int(s)}s" if daily_seconds > 0 else ""
                
                # CẬP NHẬT: Dùng alt_checkins_col để update
                alt_checkins_col.update_many(
                    {"EmployeeId": emp_id, "CheckinDate": date_str, "CheckType": {"$in": ["checkin", "checkout"]}},
                    {"$set": {"DailyHours": daily_hours, "_dailySeconds": daily_seconds}}
                )
            
            monthly_groups = {}
            for (map_emp_id, map_date_str), daily_seconds in daily_hours_map.items():
                if map_emp_id == emp_id:
                    try: month_key = datetime.strptime(map_date_str, "%d/%m/%Y").strftime("%Y-%m")
                    except: continue
                    monthly_groups.setdefault(month_key, []).append((map_date_str, daily_seconds))
            
            for month, days in monthly_groups.items():
                sorted_days = sorted(days, key=lambda x: datetime.strptime(x[0], "%d/%m/%Y"))
                running_total = 0
                for date_str, daily_seconds in sorted_days:
                    running_total += daily_seconds
                    monthly_hours_map[(emp_id, date_str)] = running_total
                    h, rem = divmod(running_total, 3600)
                    m, s = divmod(rem, 60)
                    monthly_hours = f"{int(h)}h {int(m)}m {int(s)}s" if running_total > 0 else ""
                    
                    # CẬP NHẬT: Dùng alt_checkins_col để update
                    alt_checkins_col.update_many(
                        {"EmployeeId": emp_id, "CheckinDate": date_str, "CheckType": {"$in": ["checkin", "checkout"]}},
                        {"$set": {"MonthlyHours": monthly_hours, "_monthlySeconds": running_total}}
                    )
        
        for item in all_relevant_data:
            emp_id, date_str = item.get("EmployeeId"), item.get("CheckinDate")
            daily_sec = daily_hours_map.get((emp_id, date_str), 0)
            h, rem = divmod(daily_sec, 3600)
            m, s = divmod(rem, 60)
            item['DailyHours'], item['_dailySeconds'] = (f"{int(h)}h {int(m)}m {int(s)}s" if daily_sec > 0 else ""), daily_sec
            monthly_sec = monthly_hours_map.get((emp_id, date_str), 0)
            h, rem = divmod(monthly_sec, 3600)
            m, s = divmod(rem, 60)
            item['MonthlyHours'], item['_monthlySeconds'] = (f"{int(h)}h {int(m)}m {int(s)}s" if monthly_sec > 0 else ""), monthly_sec
            if item.get('Timestamp'):
                try:
                    if isinstance(item['Timestamp'], str):
                        timestamp = datetime.strptime(item['Timestamp'], "%Y-%m-%d %H:%M:%S")
                    elif isinstance(item['Timestamp'], datetime):
                        timestamp = item['Timestamp']
                    else:
                        timestamp = None
                    item['CheckinTime'] = timestamp.astimezone(VN_TZ).strftime('%H:%M:%S') if timestamp else ""
                except (ValueError, TypeError):
                    item['CheckinTime'] = ""
        return jsonify(all_relevant_data)
    except Exception as e:
        print(f"❌ Lỗi tại get_attendances: {e}")
        return jsonify({"error": str(e)}), 500

# ---- API lấy dữ liệu nghỉ phép ----
@app.route("/api/leaves", methods=["GET"])
def get_leaves():
    try:
        email = request.args.get("email")
        # CẬP NHẬT: Dùng get_collection
        admin = get_collection("admins").find_one({"email": email})
        user = get_collection("users").find_one({"email": email})
        if not admin and not user: return jsonify({"error": "🚫 Email không tồn tại"}), 403
        
        username = None if admin else user["username"]
        date_type = request.args.get("dateType", "CheckinDate")
        filter_type = request.args.get("filter", "tất cả").lower()
        start_date_str = request.args.get("startDate")
        end_date_str = request.args.get("endDate")
        search = request.args.get("search", "").strip()

        query = build_leave_query(filter_type, start_date_str, end_date_str, search, date_type, username=username)
        
        # CẬP NHẬT: Dùng get_collection("alt_checkins")
        data = list(get_collection("alt_checkins").find(query, {"_id": 0}))

        # ... (Logic xử lý date_type == "LeaveDate" giữ nguyên) ...
        if date_type == "LeaveDate" and filter_type != "tất cả":
            filter_start_dt, filter_end_dt = None, None
            today = datetime.now(VN_TZ)
            if filter_type == "custom" and start_date_str and end_date_str:
                try:
                    filter_start_dt = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                    filter_end_dt = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                except (ValueError, TypeError): pass
            else:
                if filter_type == "hôm nay": filter_start_dt = filter_end_dt = today.date()
                elif filter_type == "tuần":
                    filter_start_dt = (today - timedelta(days=today.weekday())).date()
                    filter_end_dt = (filter_start_dt + timedelta(days=6))
                elif filter_type == "tháng":
                    filter_start_dt = today.replace(day=1).date()
                    _, last_day = calendar.monthrange(today.year, today.month)
                    filter_end_dt = today.replace(day=last_day).date()
                elif filter_type == "năm":
                    filter_start_dt = today.replace(month=1, day=1).date()
                    filter_end_dt = today.replace(month=12, day=31).date()

            if filter_start_dt and filter_end_dt:
                filtered_data = []
                for item in data:
                    display_date = item.get("DisplayDate", "")
                    if not display_date: continue
                    record_start_dt, record_end_dt = None, None
                    try:
                        if "đến" in display_date:
                            dates = re.findall(r"\d{4}-\d{2}-\d{2}", display_date)
                            if len(dates) == 2:
                                record_start_dt = datetime.strptime(dates[0], "%Y-%m-%d").date()
                                record_end_dt = datetime.strptime(dates[1], "%Y-%m-%d").date()
                        else:
                            date_part = display_date.split()[0]
                            record_start_dt = record_end_dt = datetime.strptime(date_part, "%Y-%m-%d").date()
                        
                        if record_start_dt and record_end_dt:
                            if record_start_dt <= filter_end_dt and record_end_dt >= filter_start_dt:
                                filtered_data.append(item)
                    except (ValueError, TypeError, IndexError): continue
                data = filtered_data

        if not data:
            return jsonify([])

        for item in data:
            item["ApprovalDate1"] = get_formatted_approval_date(item.get("ApprovalDate1"))
            item["ApprovalDate2"] = get_formatted_approval_date(item.get("ApprovalDate2"))
            item["Status1"] = item.get("Status1", "")
            item["Status2"] = item.get("Status2", "")
            item["Note"] = item.get("LeaveNote", "")
            if item.get('CreationTime'):
                try:
                    timestamp = item['CreationTime'] if isinstance(item['CreationTime'], datetime) else datetime.fromisoformat(item['CreationTime'].replace('Z', '+00:00'))
                    item['CheckinTime'] = timestamp.astimezone(VN_TZ).strftime('%d/%m/%Y %H:%M:%S')
                except (ValueError, TypeError):
                    item['CheckinTime'] = ""
            else:
                item['CheckinTime'] = ""
            
            display_date = item.get('DisplayDate', "")
            if display_date:
                def reformat_date(match):
                    return datetime.strptime(match.group(0), "%Y-%m-%d").strftime("%d/%m/%Y")
                display_date = re.sub(r"\d{4}-\d{2}-\d{2}", reformat_date, display_date)
            
            item['CheckinDate'] = display_date
            tasks = item.get("Tasks", [])
            tasks_str = (", ".join(tasks) if isinstance(tasks, list) else str(tasks or "")).replace("Nghỉ phép: ", "")
            item['Tasks'] = item.get("Reason") or tasks_str

        return jsonify(data)
    except Exception as e:
        import traceback
        print(f"❌ Lỗi tại get_leaves: {e}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

def get_export_date_range():
    # ... (Giữ nguyên hàm này) ...
    start_date = request.args.get("startDate")
    end_date = request.args.get("endDate")
    month = request.args.get("month")
    year = request.args.get("year")
    if month and year:
        try:
            month_int = int(month)
            year_int = int(year)
            start_dt = datetime(year_int, month_int, 1)
            _, last_day = calendar.monthrange(year_int, month_int)
            end_dt = datetime(year_int, month_int, last_day)
            return start_dt.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d")
        except (ValueError, OverflowError): pass
    if start_date and end_date:
        try:
            datetime.strptime(start_date, "%Y-%m-%d")
            datetime.strptime(end_date, "%Y-%m-%d")
            return start_date, end_date
        except ValueError: pass
    return None, None

def get_export_filename(prefix, start_date, end_date, export_date_str):
    # ... (Giữ nguyên hàm này) ...
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        start_str = start_dt.strftime("%d-%m-%Y")
        end_str = end_dt.strftime("%d-%m-%Y")
        if start_dt == end_dt: file_prefix = start_str
        elif start_dt.replace(day=1) == end_dt.replace(day=1) and start_dt.day == 1 and end_dt.day == (end_dt.replace(day=28) + timedelta(days=4)).day:
            file_prefix = f"Tháng {end_dt.month:02d}-{end_dt.year}"
        else: file_prefix = f"{start_str} đến {end_str}"
        return f"{prefix} {file_prefix}_{export_date_str}.xlsx"
    except: return f"{prefix} {start_date} đến {end_date}_{export_date_str}.xlsx"

def add_leave_date_filter(conditions, start_date, end_date):
    # ... (Giữ nguyên hàm này) ...
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        query_start = start_dt.replace(day=1).strftime("%Y-%m-%d")
        conditions.append({
            "$or": [
                {"LeaveDate": {"$gte": query_start, "$lte": end_date}},
                {"StartDate": {"$gte": query_start, "$lte": end_date}},
                {"EndDate": {"$gte": query_start, "$lte": end_date}},
                {"$and": [{"StartDate": {"$lte": end_date}}, {"EndDate": {"$gte": query_start}}]}
            ]
        })
    except: pass

def is_leave_in_range(display_date, start_dt, end_dt):
    # ... (Giữ nguyên hàm này) ...
    if not display_date: return False
    display_date = display_date.strip()
    record_start = record_end = None
    match_single = re.match(r"(\d{4}-\d{2}-\d{2})", display_date)
    if match_single:
        try: record_start = record_end = datetime.strptime(match_single.group(1), "%Y-%m-%d")
        except: return False
    else:
        match_range = re.match(r"Từ\s+(\d{4}-\d{2}-\d{2})\s+đến\s+(\d{4}-\d{2}-\d{2})", display_date)
        if match_range:
            try:
                record_start = datetime.strptime(match_range.group(1), "%Y-%m-%d")
                record_end = datetime.strptime(match_range.group(2), "%Y-%m-%d")
            except: return False
    if not record_start or not record_end: return False
    return record_start <= end_dt and record_end >= start_dt

@app.route("/api/export-excel", methods=["GET"])
def export_to_excel():
    try:
        email = request.args.get("email")
        if not email: return jsonify({"error": "Thiếu email"}), 400
        # CẬP NHẬT: Dùng get_collection
        admin = get_collection("admins").find_one({"email": email})
        user = get_collection("users").find_one({"email": email})
        if not admin and not user: return jsonify({"error": "Email không tồn tại"}), 403
        username = None if admin else user.get("username")

        start_date, end_date = get_export_date_range()
        if not start_date or not end_date: return jsonify({"error": "Thiếu thông tin ngày xuất"}), 400

        try:
            start_dt_for_query = datetime.strptime(start_date, "%Y-%m-%d")
            query_start = start_dt_for_query.replace(day=1).strftime("%Y-%m-%d")
        except Exception as e:
            print(f"Lỗi khi parse start_date: {e}")
            return jsonify({"error": "Lỗi định dạng ngày"}), 400
        
        search = request.args.get("search", "").strip()
        attendance_query = build_attendance_query("custom", query_start, end_date, search, username=username)
        # CẬP NHẬT: Dùng get_collection("alt_checkins")
        data = list(get_collection("alt_checkins").find(attendance_query, {"_id": 0}))

        template_path = "templates/Copy of Form chấm công.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def get_vn_date_str(rec):
            ts = rec.get("Timestamp")
            if not ts: return None
            try:
                if isinstance(ts, str): ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                ts_vn = ts.astimezone(VN_TZ)
                return ts_vn.strftime("%d/%m/%Y")
            except Exception: return None

        emp_data = {}
        for rec in data:
            emp_id = rec.get("EmployeeId")
            if emp_id:
                rec["_vn_date_str"] = get_vn_date_str(rec)
                if rec["_vn_date_str"]: emp_data.setdefault(emp_id, []).append(rec)

        daily_hours_map = {}
        for emp_id, records in emp_data.items():
            daily_groups = {}
            for rec in records:
                date_str = rec.get("_vn_date_str") 
                if date_str: daily_groups.setdefault(date_str, []).append(rec)

            for date_str, day_records in daily_groups.items():
                checkins = []
                checkouts = []
                for r in day_records:
                    ts = r.get("Timestamp")
                    if not ts: continue
                    try:
                        if isinstance(ts, str): ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                        if r.get("CheckType") == "checkin": checkins.append(ts)
                        elif r.get("CheckType") == "checkout": checkouts.append(ts)
                    except: continue
                
                daily_seconds = 0
                if checkins and checkouts:
                    first_in = min(checkins)
                    last_out = max(checkouts)
                    if last_out > first_in: daily_seconds = (last_out - first_in).total_seconds()
                daily_hours_map[(emp_id, date_str)] = daily_seconds

        monthly_hours_map = {}
        for emp_id in emp_data.keys():
            emp_daily_totals = { date_str: secs for (e_id, date_str), secs in daily_hours_map.items() if e_id == emp_id }
            monthly_groups = {}
            for date_str, secs in emp_daily_totals.items():
                try:
                    month_key = datetime.strptime(date_str, "%d/%m/%Y").strftime("%Y-%m")
                    monthly_groups.setdefault(month_key, []).append((date_str, secs))
                except ValueError: continue

            sorted_month_keys = sorted(monthly_groups.keys())
            for month_key in sorted_month_keys:
                month_days = monthly_groups[month_key]
                sorted_days = sorted(month_days, key=lambda x: datetime.strptime(x[0], "%d/%m/%Y"))
                running_total = 0
                for date_str, daily_sec in sorted_days:
                    running_total += daily_sec
                    h, rem = divmod(running_total, 3600)
                    m, s = divmod(rem, 60)
                    monthly_hours_map[(emp_id, date_str)] = f"{int(h)}h {int(m)}m {int(s)}s" if running_total > 0 else ""

        grouped = {}
        for d in data:
            date_str = d.get("_vn_date_str", "")
            if not date_str: continue
            try:
                rec_date = datetime.strptime(date_str, "%d/%m/%Y")
                start_dt_obj = datetime.strptime(start_date, "%Y-%m-%d")
                end_dt_obj = datetime.strptime(end_date, "%Y-%m-%d")
                if start_dt_obj.date() <= rec_date.date() <= end_dt_obj.date():
                    key = (d.get("EmployeeId", ""), d.get("EmployeeName", ""), date_str)
                    grouped.setdefault(key, []).append(d)
            except: continue

        start_row = 2
        for i, ((emp_id, emp_name, date_str), records) in enumerate(grouped.items()):
            row = start_row + i
            ws.cell(row=row, column=1, value=emp_id)
            ws.cell(row=row, column=2, value=emp_name)
            ws.cell(row=row, column=3, value=date_str) 

            daily_sec = daily_hours_map.get((emp_id, date_str), 0)
            h_d, rem_d = divmod(daily_sec, 3600)
            m_d, s_d = divmod(rem_d, 60)
            ws.cell(row=row, column=14, value=f"{int(h_d)}h {int(m_d)}m {int(s_d)}s" if daily_sec > 0 else "")
            ws.cell(row=row, column=15, value=monthly_hours_map.get((emp_id, date_str), ""))

            checkin_counter = 0
            for rec in sorted(records, key=lambda x: x.get("Timestamp") or datetime.min):
                if rec.get("CheckType") == "checkin" and checkin_counter < 9:
                    time_str = ""
                    if rec.get("Timestamp"):
                        try:
                            ts = rec["Timestamp"]
                            if isinstance(ts, str): ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                            time_str = ts.astimezone(VN_TZ).strftime("%H:%M:%S")
                        except: time_str = ""
                    tasks_str = ", ".join(rec.get("Tasks", [])) if isinstance(rec.get("Tasks"), list) else str(rec.get("Tasks", ""))
                    cell_value = "; ".join(filter(None, [time_str, rec.get("ProjectId", ""), tasks_str, rec.get("Address", ""), rec.get("CheckinNote", "")]))
                    ws.cell(row=row, column=4 + checkin_counter, value=cell_value)
                    checkin_counter += 1
                elif rec.get("CheckType") == "checkout":
                    time_str = ""
                    if rec.get("Timestamp"):
                        try:
                            ts = rec["Timestamp"]
                            if isinstance(ts, str): ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                            time_str = ts.astimezone(VN_TZ).strftime("%H:%M:%S")
                        except: time_str = ""
                    tasks_str = ", ".join(rec.get("Tasks", [])) if isinstance(rec.get("Tasks"), list) else str(rec.get("Tasks", ""))
                    cell_value = "; ".join(filter(None, [time_str, rec.get("ProjectId", ""), tasks_str, rec.get("Address", ""), rec.get("CheckinNote", "")]))
                    ws.cell(row=row, column=13, value=cell_value)

            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = align_left

        export_date_str = datetime.now(VN_TZ).strftime('%d-%m-%Y')
        filename = get_export_filename("Chấm công", start_date, end_date, export_date_str)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print(f"Lỗi export: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/api/export-leaves-excel", methods=["GET"])
def export_leaves_to_excel():
    try:
        email = request.args.get("email")
        if not email: return jsonify({"error": "Thiếu email"}), 400
        # CẬP NHẬT: Dùng get_collection
        admin = get_collection("admins").find_one({"email": email})
        user_doc = get_collection("users").find_one({"email": email})
        username = None if admin else user_doc.get("username") if user_doc else None
        if not admin and not user_doc: return jsonify({"error": "Email không tồn tại"}), 403

        start_date, end_date = get_export_date_range()
        if not start_date or not end_date: return jsonify({"error": "Thiếu thông tin ngày xuất"}), 400

        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            export_year = start_dt.year
            export_month = start_dt.month
        except:
            today = datetime.now(VN_TZ)
            start_dt = end_dt = today
            export_year = today.year
            export_month = today.month

        search = request.args.get("search", "").strip()
        base_conditions = [{"DisplayDate": {"$exists": True, "$ne": ""}}]
        if search:
            regex = re.compile(search, re.IGNORECASE)
            base_conditions.append({"$or": [{"EmployeeId": regex}, {"EmployeeName": regex}]})
        if username:
            base_conditions.append({"EmployeeName": username})

        query = {"$and": base_conditions}
        # CẬP NHẬT: Dùng get_collection("alt_checkins")
        all_leaves_data = list(get_collection("alt_checkins").find(query, {"_id": 0}))

        filtered_leaves = []
        for rec in all_leaves_data:
            display_date = rec.get("DisplayDate", "").strip()
            if is_leave_in_range(display_date, start_dt, end_dt):
                filtered_leaves.append(rec)

        template_path = "templates/Copy of Form nghỉ phép.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active

        headers = ["Mã NV", "Tên NV", "Ngày Nghỉ", "Số ngày nghỉ", "Ngày tạo đơn", "Lý do",
                   "Ngày Duyệt/Từ chối", "Trạng thái", "Ghi chú"]
        for i, h in enumerate(headers, 1): ws.cell(row=1, column=i, value=h)

        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        current_row = 2

        for rec in filtered_leaves:
            display_date_raw = rec.get("DisplayDate", "").strip()
            leave_days, is_overlap = calculate_leave_days_for_month(rec, export_year, export_month)
            def reformat_date(match): return datetime.strptime(match.group(0), "%Y-%m-%d").strftime("%d/%m/%Y")
            display_date = re.sub(r"\d{4}-\d{2}-\d{2}", reformat_date, display_date_raw)

            ws.cell(row=current_row, column=1, value=rec.get("EmployeeId", ""))
            ws.cell(row=current_row, column=2, value=rec.get("EmployeeName", ""))
            ws.cell(row=current_row, column=3, value=display_date)
            ws.cell(row=current_row, column=4, value=leave_days if is_overlap else 0)

            timestamp_str = ""
            if rec.get("CreationTime"):
                try:
                    ct = rec['CreationTime']
                    dt = ct if isinstance(ct, datetime) else datetime.fromisoformat(ct.replace('Z', '+00:00'))
                    timestamp_str = dt.astimezone(VN_TZ).strftime('%d/%m/%Y %H:%M:%S')
                except: timestamp_str = str(rec.get("CreationTime"))
            ws.cell(row=current_row, column=5, value=timestamp_str)

            tasks_str = (", ".join(rec.get("Tasks", [])) if isinstance(rec.get("Tasks"), list) else str(rec.get("Tasks", ""))).replace("Nghỉ phép: ", "")
            ws.cell(row=current_row, column=6, value=rec.get("Reason") or tasks_str)
            ws.cell(row=current_row, column=7, value=get_formatted_approval_date(rec.get("ApprovalDate2")))
            ws.cell(row=current_row, column=8, value=rec.get("Status2", ""))
            ws.cell(row=current_row, column=9, value=rec.get("LeaveNote", ""))

            for col in range(1, 10):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                cell.alignment = align_left
            current_row += 1

        export_date_str = datetime.now(VN_TZ).strftime('%d-%m-%Y')
        filename = get_export_filename("Nghỉ phép", start_date, end_date, export_date_str)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print(f"Lỗi export leaves: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/api/export-combined-excel", methods=["GET"])
def export_combined_to_excel():
    try:
        email = request.args.get("email")
        if not email: return jsonify({"error": "Thiếu email"}), 400
        # CẬP NHẬT: Dùng get_collection
        admin = get_collection("admins").find_one({"email": email})
        user = get_collection("users").find_one({"email": email})
        if not admin and not user: return jsonify({"error": "Email không tồn tại"}), 403
        username = None if admin else user.get("username")

        start_date, end_date = get_export_date_range()
        if not start_date or not end_date: return jsonify({"error": "Thiếu thông tin ngày xuất"}), 400

        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        except:
            today = datetime.now(VN_TZ)
            start_dt = end_dt = today

        try:
            start_dt_for_query = datetime.strptime(start_date, "%Y-%m-%d")
            query_start = start_dt_for_query.replace(day=1).strftime("%Y-%m-%d")
        except Exception as e:
            print(f"Lỗi khi parse start_date: {e}")
            return jsonify({"error": "Lỗi định dạng ngày"}), 400
        
        search = request.args.get("search", "").strip()
        attendance_query = build_attendance_query("custom", query_start, end_date, search, username=username)
        # CẬP NHẬT: Dùng get_collection("alt_checkins")
        attendance_data = list(get_collection("alt_checkins").find(attendance_query, {"_id": 0}))

        base_conditions = [{"DisplayDate": {"$exists": True, "$ne": ""}}]
        if search:
            regex = re.compile(search, re.IGNORECASE)
            base_conditions.append({"$or": [{"EmployeeId": regex}, {"EmployeeName": regex}]})
        if username: base_conditions.append({"EmployeeName": username})

        leave_query = {"$and": base_conditions}
        # CẬP NHẬT: Dùng get_collection("alt_checkins")
        all_leave_data = list(get_collection("alt_checkins").find(leave_query, {"_id": 0}))

        leave_data = []
        for rec in all_leave_data:
            display_date = rec.get("DisplayDate", "").strip()
            if is_leave_in_range(display_date, start_dt, end_dt): leave_data.append(rec)

        template_path = "templates/Form kết hợp.xlsx"
        wb = load_workbook(template_path)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # ================= SHEET: ĐIỂM DANH =================
        ws_att = wb["Điểm danh"]
        def get_vn_date_str_att(rec):
            ts = rec.get("Timestamp")
            if not ts: return None
            try:
                if isinstance(ts, str): ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                ts_vn = ts.astimezone(VN_TZ)
                return ts_vn.strftime("%d/%m/%Y")
            except Exception: return None

        daily_hours_map = {}
        emp_data = {}
        for rec in attendance_data:
            emp_id = rec.get("EmployeeId")
            if emp_id:
                rec["_vn_date_str"] = get_vn_date_str_att(rec)
                if rec["_vn_date_str"]: emp_data.setdefault(emp_id, []).append(rec)

        for emp_id, records in emp_data.items():
            daily_groups = {}
            for rec in records:
                date_str = rec.get("_vn_date_str")
                if date_str: daily_groups.setdefault(date_str, []).append(rec)

            for date_str, day_records in daily_groups.items():
                checkins = []
                checkouts = []
                for r in day_records:
                    ts = r.get("Timestamp")
                    if not ts: continue
                    try:
                        if isinstance(ts, str): ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                        if r.get("CheckType") == "checkin": checkins.append(ts)
                        elif r.get("CheckType") == "checkout": checkouts.append(ts)
                    except: continue
                daily_seconds = 0
                if checkins and checkouts:
                    first_in = min(checkins)
                    last_out = max(checkouts)
                    if last_out > first_in: daily_seconds = (last_out - first_in).total_seconds()
                daily_hours_map[(emp_id, date_str)] = daily_seconds

        monthly_hours_map = {}
        for emp_id in emp_data.keys():
            emp_daily_totals = { date_str: secs for (e_id, date_str), secs in daily_hours_map.items() if e_id == emp_id }
            monthly_groups = {}
            for date_str, secs in emp_daily_totals.items():
                try:
                    month_key = datetime.strptime(date_str, "%d/%m/%Y").strftime("%Y-%m")
                    monthly_groups.setdefault(month_key, []).append((date_str, secs))
                except ValueError: continue
            sorted_month_keys = sorted(monthly_groups.keys())
            for month_key in sorted_month_keys:
                month_days = monthly_groups[month_key]
                sorted_days = sorted(month_days, key=lambda x: datetime.strptime(x[0], "%d/%m/%Y"))
                running_total = 0
                for date_str, daily_sec in sorted_days:
                    running_total += daily_sec
                    h, rem = divmod(running_total, 3600)
                    m, s = divmod(rem, 60)
                    monthly_hours_map[(emp_id, date_str)] = f"{int(h)}h {int(m)}m {int(s)}s" if running_total > 0 else ""

        grouped = {}
        for d in attendance_data:
            date_str = d.get("_vn_date_str", "")
            if not date_str: continue
            try:
                rec_date = datetime.strptime(date_str, "%d/%m/%Y")
                if start_dt.date() <= rec_date.date() <= end_dt.date():
                    key = (d.get("EmployeeId", ""), d.get("EmployeeName", ""), date_str)
                    grouped.setdefault(key, []).append(d)
            except: continue

        start_row = 2
        for i, ((emp_id, emp_name, date_str), records) in enumerate(grouped.items()):
            row = start_row + i
            ws_att.cell(row=row, column=1, value=emp_id)
            ws_att.cell(row=row, column=2, value=emp_name)
            ws_att.cell(row=row, column=3, value=date_str)
            daily_sec = daily_hours_map.get((emp_id, date_str), 0)
            h_d, rem_d = divmod(daily_sec, 3600)
            m_d, s_d = divmod(rem_d, 60)
            ws_att.cell(row=row, column=14, value=f"{int(h_d)}h {int(m_d)}m {int(s_d)}s" if daily_sec > 0 else "")
            ws_att.cell(row=row, column=15, value=monthly_hours_map.get((emp_id, date_str), ""))
            checkin_counter = 0
            for rec in sorted(records, key=lambda x: x.get("Timestamp") or datetime.min):
                if rec.get("CheckType") == "checkin" and checkin_counter < 9:
                    time_str = ""
                    if rec.get("Timestamp"):
                        try:
                            ts = rec["Timestamp"]
                            if isinstance(ts, str): ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                            time_str = ts.astimezone(VN_TZ).strftime("%H:%M:%S")
                        except: time_str = ""
                    tasks_str = ", ".join(rec.get("Tasks", [])) if isinstance(rec.get("Tasks"), list) else str(rec.get("Tasks", ""))
                    cell_value = "; ".join(filter(None, [time_str, rec.get("ProjectId", ""), tasks_str, rec.get("Address", ""), rec.get("CheckinNote", "")]))
                    ws_att.cell(row=row, column=4 + checkin_counter, value=cell_value)
                    checkin_counter += 1
                elif rec.get("CheckType") == "checkout":
                    time_str = ""
                    if rec.get("Timestamp"):
                        try:
                            ts = rec["Timestamp"]
                            if isinstance(ts, str): ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                            time_str = ts.astimezone(VN_TZ).strftime("%H:%M:%S")
                        except: time_str = ""
                    tasks_str = ", ".join(rec.get("Tasks", [])) if isinstance(rec.get("Tasks"), list) else str(rec.get("Tasks", ""))
                    cell_value = "; ".join(filter(None, [time_str, rec.get("ProjectId", ""), tasks_str, rec.get("Address", ""), rec.get("CheckinNote", "")]))
                    ws_att.cell(row=row, column=13, value=cell_value)
            for col in range(1, 16):
                cell = ws_att.cell(row=row, column=col)
                cell.border = border
                cell.alignment = align_left

        # ================= SHEET: NGHỈ PHÉP =================
        ws_leave = wb["Nghỉ phép"]
        headers = ["Mã NV", "Tên NV", "Ngày Nghỉ", "Số ngày nghỉ", "Ngày tạo đơn", "Lý do", "Ngày Duyệt/Từ chối", "Trạng thái", "Ghi chú"]
        for i, h in enumerate(headers, 1): ws_leave.cell(row=1, column=i, value=h)
        current_row_leave = 2
        export_year = end_dt.year
        export_month = end_dt.month
        for rec in leave_data:
            display_date_raw = rec.get("DisplayDate", "").strip()
            leave_days, is_overlap = calculate_leave_days_for_month(rec, export_year, export_month)
            def reformat_date(match): return datetime.strptime(match.group(0), "%Y-%m-%d").strftime("%d/%m/%Y")
            display_date = re.sub(r"\d{4}-\d{2}-\d{2}", reformat_date, display_date_raw)
            ws_leave.cell(row=current_row_leave, column=1, value=rec.get("EmployeeId", ""))
            ws_leave.cell(row=current_row_leave, column=2, value=rec.get("EmployeeName", ""))
            ws_leave.cell(row=current_row_leave, column=3, value=display_date)
            ws_leave.cell(row=current_row_leave, column=4, value=leave_days if is_overlap else 0)
            timestamp_str = ""
            if rec.get("CreationTime"):
                try:
                    ct = rec['CreationTime']
                    dt = ct if isinstance(ct, datetime) else datetime.fromisoformat(ct.replace('Z', '+00:00'))
                    timestamp_str = dt.astimezone(VN_TZ).strftime('%d/%m/%Y %H:%M:%S')
                except: timestamp_str = str(rec.get("CreationTime"))
            ws_leave.cell(row=current_row_leave, column=5, value=timestamp_str)
            tasks_str = (", ".join(rec.get("Tasks", [])) if isinstance(rec.get("Tasks"), list) else str(rec.get("Tasks", ""))).replace("Nghỉ phép: ", "")
            ws_leave.cell(row=current_row_leave, column=6, value=rec.get("Reason") or tasks_str)
            ws_leave.cell(row=current_row_leave, column=7, value=get_formatted_approval_date(rec.get("ApprovalDate2")))
            ws_leave.cell(row=current_row_leave, column=8, value=rec.get("Status2", ""))
            ws_leave.cell(row=current_row_leave, column=9, value=rec.get("LeaveNote", ""))
            for col in range(1, 10):
                cell = ws_leave.cell(row=current_row_leave, column=col)
                cell.border = border
                cell.alignment = align_left
            current_row_leave += 1

        export_date_str = datetime.now(VN_TZ).strftime('%d-%m-%Y')
        filename = get_export_filename("Báo cáo tổng hợp", start_date, end_date, export_date_str)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print(f"Lỗi export combined: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# if __name__ == "__main__":
#     app.run(host="0.0.0.0", port=5000, debug=False)

