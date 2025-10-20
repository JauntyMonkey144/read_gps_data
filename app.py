import os
import re
import calendar
import secrets
import smtplib
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
from io import BytesIO
from collections import defaultdict

from dotenv import load_dotenv
from flask import (Flask, jsonify, redirect, render_template, request,
                   send_file, url_for)
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from pymongo import MongoClient
from werkzeug.security import check_password_hash, generate_password_hash

# ---- Cấu hình ứng dụng ----
app = Flask(__name__, template_folder="templates")
CORS(app, methods=["GET", "POST"])

# ---- Tải biến môi trường ----
load_dotenv()

# ---- Timezone Việt Nam ----
VN_TZ = timezone(timedelta(hours=7))

# ---- Cấu hình MongoDB ----
MONGO_URI = os.getenv("MONGO_URI")
DB_NAME = os.getenv("DB_NAME", "Sun_Database_1")
client = MongoClient(MONGO_URI)
db = client[DB_NAME]

# ---- Cấu hình Email ----
EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))

# ---- Collections ----
admins = db["admins"]
users = db["users"]
collection = db["alt_checkins"]
reset_tokens = db["reset_tokens"]

# ==============================================================================
# ---- CÁC ROUTE XÁC THỰC VÀ QUẢN LÝ TÀI KHOẢN ----
# ==============================================================================

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/login", methods=["POST", "GET"])
def login():
    if request.method == "GET":
        return redirect(url_for("index"))

    email = request.form.get("email")
    password = request.form.get("password")
    if not email or not password:
        return jsonify({"success": False, "message": "❌ Vui lòng nhập email và mật khẩu"}), 400

    admin = admins.find_one({"email": email})
    if admin and check_password_hash(admin.get("password", ""), password):
        return jsonify({
            "success": True, "message": "✅ Đăng nhập thành công",
            "username": admin["username"], "email": admin["email"], "role": "admin"
        })

    user = users.find_one({"email": email})
    if user and check_password_hash(user.get("password", ""), password):
        return jsonify({
            "success": True, "message": "✅ Đăng nhập thành công",
            "username": user["username"], "email": user["email"], "role": "user"
        })

    return jsonify({"success": False, "message": "🚫 Email hoặc mật khẩu không đúng!"}), 401

@app.route("/forgot-password", methods=["GET"])
def forgot_password():
    return render_template("forgot_password.html")

@app.route("/request-reset-password", methods=["POST"])
def request_reset_password():
    email = request.form.get("email")
    if not email:
        return render_template("message.html", message="❌ Vui lòng nhập email", is_error=True), 400

    account = admins.find_one({"email": email}) or users.find_one({"email": email})
    if not account:
        return render_template("message.html", message="🚫 Email không tồn tại!", is_error=True), 404

    token = secrets.token_urlsafe(32)
    expiration = datetime.now(timezone.utc) + timedelta(hours=1)
    reset_tokens.insert_one({"email": email, "token": token, "expiration": expiration})

    try:
        reset_link = url_for("reset_password_form", token=token, _external=True)
        msg = MIMEMultipart()
        msg['From'] = formataddr(("Sun Automation System", EMAIL_ADDRESS))
        msg['To'] = email
        msg['Subject'] = "Yêu cầu đặt lại mật khẩu"
        
        html_body = render_template("email_reset.html", reset_link=reset_link)
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(msg)

        return render_template("message.html", message="✅ Email chứa liên kết đặt lại mật khẩu đã được gửi! Vui lòng kiểm tra hộp thư của bạn.", is_error=False)
    except Exception as e:
        app.logger.error(f"❌ Lỗi gửi email: {e}")
        return render_template("message.html", message="❌ Lỗi khi gửi email, vui lòng thử lại sau.", is_error=True), 500

@app.route("/reset-password/<token>", methods=["GET"])
def reset_password_form(token):
    token_data = reset_tokens.find_one({"token": token})
    if not token_data or token_data["expiration"] < datetime.now(timezone.utc):
        return render_template("message.html", message="🚫 Liên kết không hợp lệ hoặc đã hết hạn!", is_error=True), 400
    return render_template("reset_password.html", token=token)

@app.route("/update-password", methods=["POST"])
def update_password():
    token = request.form.get("token")
    new_password = request.form.get("new_password")
    confirm_password = request.form.get("confirm_password")

    token_data = reset_tokens.find_one({"token": token})
    if not token_data or token_data["expiration"] < datetime.now(timezone.utc):
        return render_template("message.html", message="🚫 Liên kết không hợp lệ hoặc đã hết hạn!", is_error=True), 400

    if not new_password or new_password != confirm_password:
        return render_template("message.html", message="❌ Mật khẩu không khớp hoặc không được để trống!", is_error=True, back_url=url_for('reset_password_form', token=token)), 400
    
    email = token_data["email"]
    hashed_pw = generate_password_hash(new_password)
    
    if admins.find_one({"email": email}):
        collection_to_update = admins
    else:
        collection_to_update = users

    collection_to_update.update_one({"email": email}, {"$set": {"password": hashed_pw}})
    reset_tokens.delete_one({"token": token})

    return render_template("message.html", message="✅ Mật khẩu đã được cập nhật thành công!", is_error=False, back_url=url_for('index'))

# ==============================================================================
# ---- CÁC HÀM HELPER VÀ XÂY DỰNG QUERY ----
# ==============================================================================

def build_attendance_query(filter_type, start_date, end_date, search, username=None):
    today = datetime.now(VN_TZ)
    conditions = [{"CheckType": {"$in": ["checkin", "checkout"]}}]
    date_filter = {}

    try:
        if filter_type == "custom" and start_date and end_date:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=VN_TZ)
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=VN_TZ)
            date_filter = {"Timestamp": {"$gte": start_dt, "$lte": end_dt}}
        elif filter_type == "hôm nay":
            start_dt, end_dt = today.replace(hour=0, minute=0, second=0), today.replace(hour=23, minute=59, second=59)
            date_filter = {"Timestamp": {"$gte": start_dt, "$lte": end_dt}}
        elif filter_type == "tuần":
            start_dt = (today - timedelta(days=today.weekday())).replace(hour=0, minute=0, second=0)
            end_dt = (start_dt + timedelta(days=6)).replace(hour=23, minute=59, second=59)
            date_filter = {"Timestamp": {"$gte": start_dt, "$lte": end_dt}}
        elif filter_type == "tháng":
            start_dt = today.replace(day=1, hour=0, minute=0, second=0)
            _, last_day = calendar.monthrange(today.year, today.month)
            end_dt = today.replace(day=last_day, hour=23, minute=59, second=59)
            date_filter = {"Timestamp": {"$gte": start_dt, "$lte": end_dt}}
        elif filter_type == "năm":
            start_dt = today.replace(month=1, day=1, hour=0, minute=0, second=0)
            end_dt = today.replace(month=12, day=31, hour=23, minute=59, second=59)
            date_filter = {"Timestamp": {"$gte": start_dt, "$lte": end_dt}}
    except (ValueError, TypeError):
        pass

    if date_filter:
        conditions.append(date_filter)
    if search:
        regex = re.compile(search, re.IGNORECASE)
        conditions.append({"$or": [{"EmployeeId": regex}, {"EmployeeName": regex}]})
    if username:
        conditions.append({"EmployeeName": username})
    
    return {"$and": conditions} if len(conditions) > 1 else conditions[0] if conditions else {}

def build_leave_query(filter_type, start_date_str, end_date_str, search, date_type="CheckinTime", username=None):
    today = datetime.now(VN_TZ)
    regex_leave = re.compile("Nghỉ phép", re.IGNORECASE)
    conditions = [{"$or": [{"Tasks": regex_leave}, {"Reason": {"$exists": True}}]}]
    date_filter = {}
    start_dt, end_dt = None, None

    if filter_type == "custom" and start_date_str and end_date_str:
        try:
            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d").replace(tzinfo=VN_TZ)
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=VN_TZ)
        except ValueError:
            pass
    elif filter_type != "tất cả":
        if filter_type == "hôm nay":
            start_dt, end_dt = today.replace(hour=0, minute=0, second=0), today.replace(hour=23, minute=59, second=59)
        elif filter_type == "tuần":
            start_dt = (today - timedelta(days=today.weekday())).replace(hour=0, minute=0, second=0)
            end_dt = (start_dt + timedelta(days=6)).replace(hour=23, minute=59, second=59)
        elif filter_type == "tháng":
            start_dt = today.replace(day=1, hour=0, minute=0, second=0)
            _, last_day = calendar.monthrange(today.year, today.month)
            end_dt = today.replace(day=last_day, hour=23, minute=59, second=59)
        elif filter_type == "năm":
            start_dt = today.replace(month=1, day=1, hour=0, minute=0, second=0)
            end_dt = today.replace(month=12, day=31, hour=23, minute=59, second=59)

    if start_dt and end_dt:
        if date_type == "CheckinTime":
            date_filter = {"CreationTime": {"$gte": start_dt, "$lte": end_dt}}
        elif date_type == "ApprovalDate1":
            date_filter = {"ApprovalDate1": {"$gte": start_dt, "$lte": end_dt}}
        elif date_type == "ApprovalDate2":
            date_filter = {"ApprovalDate2": {"$gte": start_dt, "$lte": end_dt}}

    if date_filter:
        conditions.append(date_filter)
    if search:
        regex = re.compile(search, re.IGNORECASE)
        conditions.append({"$or": [{"EmployeeId": regex}, {"EmployeeName": regex}]})
    if username:
        conditions.append({"EmployeeName": username})

    return {"$and": conditions}

def calculate_leave_days_for_month(record, export_year, export_month):
    display_date = record.get("DisplayDate", "").strip().lower()
    start_date, end_date = None, None
    try:
        if display_date:
            if "từ" in display_date and "đến" in display_date:
                date_parts = re.findall(r"\d{4}-\d{2}-\d{2}", display_date)
                if len(date_parts) == 2:
                    start_date = datetime.strptime(date_parts[0], "%Y-%m-%d")
                    end_date = datetime.strptime(date_parts[1], "%Y-%m-%d")
            else:
                date_part = display_date.split()[0]
                start_date = end_date = datetime.strptime(date_part, "%Y-%m-%d")
        elif record.get('StartDate') and record.get('EndDate'):
            start_date = datetime.strptime(record['StartDate'], "%Y-%m-%d")
            end_date = datetime.strptime(record['EndDate'], "%Y-%m-%d")
        elif record.get('LeaveDate'):
            start_date = end_date = datetime.strptime(record['LeaveDate'], "%Y-%m-%d")
    except (ValueError, TypeError, IndexError):
        return 0.0, False

    if not start_date or not end_date:
        return 0.0, False

    days_in_month = 0.0
    _, last_day = calendar.monthrange(export_year, export_month)
    month_start = datetime(export_year, export_month, 1)
    month_end = datetime(export_year, export_month, last_day)

    if not (start_date > month_end or end_date < month_start):
        current_date = max(start_date, month_start)
        loop_end_date = min(end_date, month_end)
        
        while current_date <= loop_end_date:
            if current_date.weekday() < 6: # 0-5 là T2-T7
                if "cả ngày" in display_date or not ("sáng" in display_date or "chiều" in display_date):
                    days_in_month += 1.0
                else:
                    days_in_month += 0.5
            current_date += timedelta(days=1)

    if days_in_month == 0:
        return 0.0, False

    status1 = str(record.get("Status1") or "").lower()
    status2 = str(record.get("Status2") or "").lower()

    if "từ chối" in status2 or ("từ chối" in status1 and not status2):
        return 0.0, True
    elif "duyệt" in status2 or "duyệt" in status1:
        return days_in_month, True
    else: # Đang chờ
        return 0.0, True

def get_formatted_approval_date(approval_date):
    if not approval_date: return ""
    try:
        if isinstance(approval_date, datetime):
            return approval_date.astimezone(VN_TZ).strftime("%d/%m/%Y %H:%M:%S")
        return str(approval_date)
    except:
        return str(approval_date)

# ==============================================================================
# ---- CÁC API TRẢ VỀ DỮ LIỆU (ĐÃ TỐI ƯU) ----
# ==============================================================================

@app.route("/api/attendances", methods=["GET"])
def get_attendances():
    try:
        email = request.args.get("email")
        admin = admins.find_one({"email": email})
        user = users.find_one({"email": email})
        if not admin and not user:
            return jsonify({"error": "🚫 Email không tồn tại"}), 403

        username = None if admin else user["username"]
        query_conditions = build_attendance_query(
            request.args.get("filter", "hôm nay").lower(),
            request.args.get("startDate"), request.args.get("endDate"),
            request.args.get("search", "").strip(), username=username
        )

        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        skip = (page - 1) * limit

        total_records = collection.count_documents(query_conditions)
        raw_records = list(collection.find(query_conditions, {"_id": 0}).sort("Timestamp", 1).skip(skip).limit(limit))
        
        if not raw_records:
            return jsonify({"data": [], "total": 0, "page": page, "totalPages": 0})
        
        daily_groups = defaultdict(lambda: {'checkins': [], 'checkouts': []})
        for rec in raw_records:
            key = (rec.get("EmployeeId"), rec.get("CheckinDate"))
            ts = rec.get('Timestamp')
            if isinstance(ts, datetime):
                if rec.get('CheckType') == 'checkin':
                    daily_groups[key]['checkins'].append(ts)
                elif rec.get('CheckType') == 'checkout':
                    daily_groups[key]['checkouts'].append(ts)
        
        daily_seconds_map = {}
        for key, value in daily_groups.items():
            checkins, checkouts = sorted(value['checkins']), sorted(value['checkouts'])
            daily_seconds = 0
            if checkins and checkouts and checkouts[-1] > checkins[0]:
                daily_seconds = (checkouts[-1] - checkins[0]).total_seconds()
            daily_seconds_map[key] = daily_seconds

        employee_data = defaultdict(list)
        for rec in raw_records:
            key = (rec.get("EmployeeId"), rec.get("CheckinDate"))
            rec['_dailySeconds'] = daily_seconds_map.get(key, 0)
            employee_data[rec.get("EmployeeId")].append(rec)
            
        final_result = []
        for emp_id, records in employee_data.items():
            sorted_records = sorted(records, key=lambda r: datetime.strptime(r['CheckinDate'], "%d/%m/%Y"))
            monthly_running_seconds = 0
            current_month = None
            for rec in sorted_records:
                record_month = rec['CheckinDate'][3:]
                if record_month != current_month:
                    monthly_running_seconds = 0
                    current_month = record_month
                
                monthly_running_seconds += rec['_dailySeconds']
                rec['_monthlySeconds'] = monthly_running_seconds
                final_result.append(rec)

        for item in final_result:
            h, rem = divmod(item.get('_dailySeconds', 0), 3600)
            m, s = divmod(rem, 60)
            item['DailyHours'] = f"{int(h)}h {int(m)}m {int(s)}s" if item.get('_dailySeconds', 0) > 0 else ""
            
            h, rem = divmod(item.get('_monthlySeconds', 0), 3600)
            m, s = divmod(rem, 60)
            item['MonthlyHours'] = f"{int(h)}h {int(m)}m {int(s)}s" if item.get('_monthlySeconds', 0) > 0 else ""
            
            ts = item.get('Timestamp')
            item['CheckinTime'] = ts.astimezone(VN_TZ).strftime('%H:%M:%S') if isinstance(ts, datetime) else ""

            item.pop('_dailySeconds', None)
            item.pop('_monthlySeconds', None)
            item.pop('_id', None)

        return jsonify({
            "data": final_result,
            "total": total_records,
            "page": page,
            "totalPages": (total_records + limit - 1) // limit
        })

    except Exception as e:
        app.logger.error(f"❌ Lỗi tại get_attendances: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/leaves", methods=["GET"])
def get_leaves():
    try:
        email = request.args.get("email")
        admin = admins.find_one({"email": email})
        user = users.find_one({"email": email})
        if not admin and not user:
            return jsonify({"error": "🚫 Email không tồn tại"}), 403

        username = None if admin else user["username"]
        date_type = request.args.get("dateType", "CheckinTime")
        filter_type = request.args.get("filter", "tất cả").lower()
        start_date_str = request.args.get("startDate")
        end_date_str = request.args.get("endDate")
        search = request.args.get("search", "").strip()

        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        skip = (page - 1) * limit

        query = build_leave_query(filter_type, start_date_str, end_date_str, search, date_type, username=username)
        
        total_records = collection.count_documents(query)
        data = list(collection.find(query, {"_id": 0}).sort("CreationTime", -1).skip(skip).limit(limit))

        for item in data:
            item["ApprovalDate1"] = get_formatted_approval_date(item.get("ApprovalDate1"))
            item["ApprovalDate2"] = get_formatted_approval_date(item.get("ApprovalDate2"))
            item["Status1"] = item.get("Status1", "")
            item["Status2"] = item.get("Status2", "")
            item["Note"] = item.get("LeaveNote", "")
            
            creation_time = item.get('CreationTime')
            if isinstance(creation_time, datetime):
                item['CheckinTime'] = creation_time.astimezone(VN_TZ).strftime('%d/%m/%Y %H:%M:%S')
            else:
                item['CheckinTime'] = ""
            
            display_date = item.get('DisplayDate', "")
            if display_date:
                def reformat_date(match):
                    return datetime.strptime(match.group(0), "%Y-%m-%d").strftime("%d/%m/%Y")
                display_date = re.sub(r"\d{4}-\d{2}-\d{2}", reformat_date, display_date)
            
            item['CheckinDate'] = display_date
            tasks = item.get("Tasks", [])
            tasks_str = (", ".join(tasks) if isinstance(tasks, list) else str(tasks or "")).replace("Nghỉ phép: ", "")
            item['Tasks'] = item.get("Reason") or tasks_str

        return jsonify({
            "data": data,
            "total": total_records,
            "page": page,
            "totalPages": (total_records + limit - 1) // limit
        })
    except Exception as e:
        app.logger.error(f"❌ Lỗi tại get_leaves: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ==============================================================================
# ---- CÁC API XUẤT FILE EXCEL ----
# ==============================================================================

@app.route("/api/export-excel", methods=["GET"])
def export_to_excel():
    try:
        email = request.args.get("email")
        admin = admins.find_one({"email": email})
        user = users.find_one({"email": email})
        if not admin and not user: return jsonify({"error": "🚫 Email không tồn tại"}), 403
        username = None if admin else user.get("username")

        export_year = int(request.args.get("year", datetime.now(VN_TZ).year))
        export_month = int(request.args.get("month", datetime.now(VN_TZ).month))
        
        start_date = datetime(export_year, export_month, 1).strftime("%Y-%m-%d")
        _, last_day = calendar.monthrange(export_year, export_month)
        end_date = datetime(export_year, export_month, last_day).strftime("%Y-%m-%d")

        query = build_attendance_query("custom", start_date, end_date, request.args.get("search", "").strip(), username=username)
        data = list(collection.find(query, {"_id": 0}))
        
        grouped = {}
        for d in data:
            key = (d.get("EmployeeId", ""), d.get("EmployeeName", ""), d.get("CheckinDate"))
            grouped.setdefault(key, []).append(d)
        
        template_path = "templates/Copy of Form chấm công.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        start_row = 2
        for i, ((emp_id, emp_name, date_str), records) in enumerate(grouped.items()):
            row = start_row + i
            ws.cell(row=row, column=1, value=emp_id)
            ws.cell(row=row, column=2, value=emp_name)
            ws.cell(row=row, column=3, value=date_str)
            
            daily_hours = records[0].get("DailyHours", "0h 0m 0s")
            monthly_hours = records[0].get("MonthlyHours", "0h 0m 0s")
            ws.cell(row=row, column=14, value=daily_hours)
            ws.cell(row=row, column=15, value=monthly_hours)
            
            checkin_counter, checkin_start_col, checkout_col = 0, 4, 13
            sorted_records = sorted(records, key=lambda x: x.get('Timestamp', datetime.min))
            
            for rec in sorted_records:
                time_str = ""
                ts = rec.get('Timestamp')
                if isinstance(ts, datetime):
                    time_str = ts.astimezone(VN_TZ).strftime("%H:%M:%S")

                tasks_str = ", ".join(rec.get("Tasks", [])) if isinstance(rec.get("Tasks"), list) else str(rec.get("Tasks", ""))
                fields = [time_str, rec.get('ProjectId', ''), tasks_str, rec.get('Address', ''), rec.get('CheckinNote', '')]
                cell_value = "; ".join(field for field in fields if field)
                
                if rec.get('CheckType') == 'checkin' and checkin_counter < 9:
                    ws.cell(row=row, column=checkin_start_col + checkin_counter, value=cell_value)
                    checkin_counter += 1
                elif rec.get('CheckType') == 'checkout':
                    ws.cell(row=row, column=checkout_col, value=cell_value)
            
            for col in range(1, 16):
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).alignment = align_left

        filename = f"Danh sách chấm công_Tháng_{export_month}-{export_year}.xlsx"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        app.logger.error(f"❌ Lỗi export excel: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/api/export-leaves-excel", methods=["GET"])
def export_leaves_to_excel():
    try:
        email = request.args.get("email")
        admin = admins.find_one({"email": email})
        user = users.find_one({"email": email})
        if not admin and not user: return jsonify({"error": "🚫 Email không tồn tại"}), 403
        username = None if admin else user.get("username")
        
        export_year = int(request.args.get("year", datetime.now(VN_TZ).year))
        export_month = int(request.args.get("month", datetime.now(VN_TZ).month))

        regex_leave = re.compile("Nghỉ phép", re.IGNORECASE)
        conditions = [{"$or": [{"Tasks": regex_leave}, {"Reason": {"$exists": True}}]}]
        search = request.args.get("search", "").strip()
        if search:
            regex = re.compile(search, re.IGNORECASE)
            conditions.append({"$or": [{"EmployeeId": regex}, {"EmployeeName": regex}]})
        if username:
            conditions.append({"EmployeeName": username})
        query = {"$and": conditions}
        all_leaves_data = list(collection.find(query, {"_id": 0}))

        template_path = "templates/Copy of Form nghỉ phép.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active
        ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'], ws['F1'], ws['G1'], ws['H1'], ws['I1'], ws['J1'], ws['K1'] = \
            "Mã NV", "Tên NV", "Ngày Nghỉ", "Số ngày nghỉ", "Ngày tạo đơn", "Lý do", "Ngày Duyệt/Từ chối Lần đầu", \
            "Trạng thái Lần đầu", "Ngày Duyệt/Từ chối Lần cuối", "Trạng thái Lần cuối", "Ghi chú"
        
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        current_row = 2
        for rec in all_leaves_data:
            leave_days, is_overlap = calculate_leave_days_for_month(rec, export_year, export_month)
            if is_overlap:
                ws.cell(row=current_row, column=1, value=rec.get("EmployeeId", ""))
                ws.cell(row=current_row, column=2, value=rec.get("EmployeeName", ""))
                ws.cell(row=current_row, column=3, value=rec.get("DisplayDate", ""))
                ws.cell(row=current_row, column=4, value=leave_days)
                ws.cell(row=current_row, column=5, value=get_formatted_approval_date(rec.get("CreationTime")))
                tasks = rec.get("Tasks", [])
                tasks_str = (", ".join(tasks) if isinstance(tasks, list) else str(tasks or "")).replace("Nghỉ phép: ", "")
                ws.cell(row=current_row, column=6, value=rec.get("Reason") or tasks_str)
                ws.cell(row=current_row, column=7, value=get_formatted_approval_date(rec.get("ApprovalDate1")))
                ws.cell(row=current_row, column=8, value=rec.get("Status1", ""))
                ws.cell(row=current_row, column=9, value=get_formatted_approval_date(rec.get("ApprovalDate2")))
                ws.cell(row=current_row, column=10, value=rec.get("Status2", ""))
                ws.cell(row=current_row, column=11, value=rec.get("LeaveNote", ""))

                for col_idx in range(1, 12):
                    ws.cell(row=current_row, column=col_idx).border = border
                    ws.cell(row=current_row, column=col_idx).alignment = align_left
                
                current_row += 1

        filename = f"Danh sách nghỉ phép_Tháng_{export_month}-{export_year}.xlsx"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        app.logger.error(f"❌ Lỗi export leaves: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ==============================================================================
# ---- CHẠY ỨNG DỤNG ----
# ==============================================================================

if __name__ == "__main__":
    # Khi deploy, hãy sử dụng một WSGI server như Gunicorn
    # Ví dụ: gunicorn --workers 4 --bind 0.0.0.0:5000 app:app
    app.run(host="0.0.0.0", port=5000, debug=True)

